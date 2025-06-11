# utils/excel_utils.py

import pandas as pd
import email.header
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def convert_csv_to_excel(csv_file_path, excel_file_path="cleaned_email_data.xlsx"):
    # Load raw CSV
    df = pd.read_csv(csv_file_path)

    # Decode MIME-encoded subject lines
    def decode_subject(subject):
        try:
            decoded, charset = email.header.decode_header(subject)[0]
            if isinstance(decoded, bytes):
                return decoded.decode(charset or "utf-8")
            return decoded
        except Exception:
            return subject

    df["Subject"] = df["Subject"].apply(decode_subject)

    # Clean whitespaces and truncate long snippets
    df["Snippet"] = df["Snippet"].fillna("").apply(lambda x: " ".join(str(x).split()))
    df["Snippet"] = df["Snippet"].apply(lambda x: x[:150] + "..." if len(x) > 150 else x)

    # Rename columns for presentation
    df = df.rename(columns={
        "Subject": "Email Subject",
        "From": "Sender",
        "Snippet": "Email Snippet",
        "AI_Category": "Predicted Category",
        "User_Category": "User Category"
    })

    # Save cleaned version to Excel
    df.to_excel(excel_file_path, index=False)

    # Auto-adjust column widths
    wb = load_workbook(excel_file_path)
    ws = wb.active
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2
    wb.save(excel_file_path)

    print(f"✅ Cleaned Excel saved as: {excel_file_path}")
